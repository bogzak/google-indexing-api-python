from openpyxl import load_workbook

file_for_inspection = load_workbook('for_url_inspection.xlsx')
file_all_urls = load_workbook('all_urls.xlsx')
sheet_for_inspection = file_for_inspection['Sheet1']
sheet_all_urls = file_all_urls['Sheet1']

def send_urls():
    row = 1
    for row_all_url in sheet_all_urls.iter_rows(max_col=1, max_row=2000):
        for cell_all_url in row_all_url:
            value_all_url = cell_all_url.value
            if value_all_url is not None:
                sheet_for_inspection.cell(row=row, column=1).value = value_all_url
            # sheet_all_urls.delete_rows(1)
            row += 1
    sheet_all_urls.delete_rows(1, 2000)
    file_for_inspection.save('for_url_inspection.xlsx')
    file_all_urls.save('all_urls.xlsx')

def main():
    send_urls()

if __name__ == '__main__':
    main()
