from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
import datetime
import httplib2
import json
import os

"""
pip install google-api-python-client oauth2client
pip install --upgrade oauth2client
pip install openpyxl
"""

SCOPES = ["https://www.googleapis.com/auth/indexing"]

def write_result(work_type, url, date):
    if work_type == 'txt_file':
        with open('result.txt', 'a', encoding='utf-8') as result_file:
            string_write = f"{url};{date}\n"
            result_file.write(string_write)


def indexURL(url, http):
    ENDPOINT = "https://indexing.googleapis.com/v3/urlNotifications:publish"
    content = {'url': url.strip(), 'type': "URL_UPDATED"}
    json_ctn = json.dumps(content)
    response, content = http.request(ENDPOINT, method="POST", body=json_ctn)
    json.loads(content.decode())

def send_urls():
    count_urls = 0
    flag = False
    for root, dirs, files in os.walk("json_keys"):
        for json_key_path_name in files:
            json_key = 'json_keys/' + json_key_path_name
            credentials = ServiceAccountCredentials.from_json_keyfile_name(json_key, scopes=SCOPES)
            http = credentials.authorize(httplib2.Http())

            url_file = load_workbook("search_console_url_is_not_on_google.xlsx")  # get file
            url_sheet = url_file["1 - URL is Not on Google"]  # get sheet
            row_num = 2
            for rows in url_sheet.iter_rows(min_row=2, max_row=201, max_col=1):
                for url in rows:
                    try:
                        url_new = str(url_sheet.cell(column=1, row=row_num).value).strip()
                        if url.value is not None:
                            indexURL(url_new, http)
                            write_result('txt_file', url_new, datetime.date.today())
                            count_urls += 1
                        else:
                            raise StopIteration
                        row_num += 1
                    except StopIteration:
                        url_sheet.delete_rows(2, row_num)
                        url_file.save("search_console_url_is_not_on_google.xlsx")
                        print("В файле больше нет страниц")
                        print("Страниц обработано: " + str(count_urls) + " шт.")
                        flag = True
                        break

                if flag:
                    break
            if flag:
                break

            url_sheet.delete_rows(2, 200)

def main():
    send_urls()

if __name__ == "__main__":
    main()
